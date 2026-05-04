from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class TestCase:
    id: str
    area: str
    title: str
    priority: str
    preconditions: str
    steps: str
    expected: str


TEST_CASES: list[TestCase] = [
    TestCase(
        id="TC-A01",
        area="Account",
        title="Register (no address field)",
        priority="High",
        preconditions="None",
        steps="1) Open Register\n2) Enter valid name/email/mobile/password\n3) Submit",
        expected="Account is created; OTP flow starts; no address-required validation.",
    ),
    TestCase(
        id="TC-A02",
        area="Account/Checkout",
        title="Login required for checkout",
        priority="High",
        preconditions="User is logged out; cart has at least 1 item",
        steps="1) Open Cart\n2) Click Place Order",
        expected="User is blocked and redirected/prompted to log in; no order is created.",
    ),
    TestCase(
        id="TC-A03",
        area="Account/Checkout",
        title="Address required for cart checkout",
        priority="High",
        preconditions="Logged-in user; profile address is empty; cart has at least 1 item",
        steps="1) Open Cart\n2) Click Place Order",
        expected="Checkout is blocked with message to set address in Account Settings; redirect option works.",
    ),
    TestCase(
        id="TC-A04",
        area="Account/Custom Design",
        title="Address required for custom design submit",
        priority="High",
        preconditions="Logged-in user; profile address is empty",
        steps="1) Open Upload Custom Design\n2) Fill required fields\n3) Click Submit Request",
        expected="Submission is blocked with message to set address in Account Settings; redirect option works.",
    ),
    TestCase(
        id="TC-A05",
        area="Account",
        title="PH address autocomplete suggestions",
        priority="Medium",
        preconditions="None",
        steps="1) Open Account Settings\n2) Type 4+ chars in Address field\n3) Pick a suggestion\n4) Press ESC / click outside",
        expected="Suggestions appear; selection fills the textarea; ESC/outside click closes suggestions.",
    ),
    TestCase(
        id="TC-B01",
        area="Cart",
        title="Checkout with GCash selected",
        priority="High",
        preconditions="Logged-in user; profile address set; cart has at least 1 item",
        steps="1) Open Cart\n2) Select GCash\n3) Place Order",
        expected="Order is created; order meta payment method is 'GCash'.",
    ),
    TestCase(
        id="TC-B02",
        area="Cart",
        title="Checkout with COD selected",
        priority="High",
        preconditions="Logged-in user; profile address set; cart has at least 1 item",
        steps="1) Open Cart\n2) Select COD\n3) Place Order",
        expected="Order is created; order meta payment method is 'COD'.",
    ),
    TestCase(
        id="TC-B03",
        area="Cart/Promo",
        title="Free shipping promo recorded for 10+ qty",
        priority="High",
        preconditions="Logged-in user; profile address set",
        steps="1) Add items so total qty >= 10\n2) Place Order\n3) Open admin order details",
        expected="Promo is tagged (free shipping 10+); admin shipping fee is enforced to 0.",
    ),
    TestCase(
        id="TC-C01",
        area="Payments",
        title="Customer uploads downpayment receipt",
        priority="High",
        preconditions="Order is in Awaiting Payment stage",
        steps="1) Open Order Tracking\n2) Choose receipt image\n3) Upload receipt",
        expected="Receipt is saved; status shows pending verification until admin verifies.",
    ),
    TestCase(
        id="TC-C02",
        area="Admin/Payments",
        title="Cannot proceed to proofing without downpayment verification",
        priority="High",
        preconditions="Order exists; downpayment receipt not verified",
        steps="1) Open Admin Order Details\n2) Attempt to move to Proofing/next stage",
        expected="Action is blocked (UI and/or server rejection).",
    ),
    TestCase(
        id="TC-C03",
        area="Admin/Payments",
        title="Admin verifies 50% downpayment (DB order)",
        priority="High",
        preconditions="Downpayment receipt uploaded; order in paid/Awaiting Payment stage",
        steps="1) Open Admin Order Details\n2) Click Verify 50% Downpayment\n3) Confirm",
        expected="Payment meta becomes verified_type=downpayment; order can proceed to Proofing.",
    ),
    TestCase(
        id="TC-C04",
        area="Admin/Payments",
        title="No 'Verify 100% Full Payment' path",
        priority="High",
        preconditions="Any order in paid/Awaiting Payment stage",
        steps="1) Open Admin Order Details\n2) Inspect available payment verification buttons",
        expected="There is no 'Verify 100% Full Payment' button; backend rejects verify_type=full.",
    ),
    TestCase(
        id="TC-D01",
        area="COD",
        title="COD skips Awaiting Final Payment flow",
        priority="High",
        preconditions="COD order; downpayment verified; order in processing",
        steps="1) Open Admin Order Details\n2) From Processing, attempt to request final payment\n3) Mark Ready to Ship",
        expected="COD does not require final receipt; Ready to Ship is allowed without final verification.",
    ),
    TestCase(
        id="TC-D02",
        area="COD",
        title="Customer cannot upload final receipt for COD",
        priority="High",
        preconditions="COD order",
        steps="1) Open Order Tracking\n2) Navigate to any final-payment UI (if visible)\n3) Try uploading final receipt",
        expected="Final receipt upload is not shown or is blocked with message: balance due on delivery.",
    ),
    TestCase(
        id="TC-D03",
        area="COD/Admin",
        title="COD completion requires 'Mark COD Final Payment Received'",
        priority="High",
        preconditions="COD order shipped; downpayment verified; final not verified",
        steps="1) Open Admin Order Details\n2) Attempt Mark Completed",
        expected="Completion is blocked until COD final payment is marked received (server-enforced).",
    ),
    TestCase(
        id="TC-D04",
        area="COD/Admin",
        title="Admin marks COD final payment received",
        priority="High",
        preconditions="COD order shipped; downpayment verified",
        steps="1) Open Admin Order Details\n2) Click 'Mark COD Final Payment Received'\n3) Confirm\n4) Mark Completed",
        expected="Payment final_verified=true and amount_paid=total; order can be completed.",
    ),
    TestCase(
        id="TC-E01",
        area="GCash",
        title="Non-COD can request final payment",
        priority="High",
        preconditions="GCash order; downpayment verified; order in processing",
        steps="1) Open Admin Order Details\n2) Click 'Set Awaiting Final Payment'",
        expected="Status updates to awaiting_final_payment; customer sees final payment instructions.",
    ),
    TestCase(
        id="TC-E02",
        area="GCash",
        title="Final receipt upload + admin verify",
        priority="High",
        preconditions="GCash order in awaiting_final_payment",
        steps="1) Customer uploads final receipt\n2) Admin verifies final payment\n3) Admin marks Ready to Ship",
        expected="Final is verified; order can proceed to shipping.",
    ),
    TestCase(
        id="TC-F01",
        area="Admin/Pricing",
        title="Admin sets shipping fee for qty < 10",
        priority="Medium",
        preconditions="Order qty < 10; order pending approval",
        steps="1) Open Admin Order Details\n2) Set Shipping Fee\n3) Save/Approve",
        expected="Shipping fee is saved; totals update; approval can proceed.",
    ),
    TestCase(
        id="TC-F02",
        area="Admin/Pricing",
        title="Admin cannot set shipping fee > 0 for qty >= 10",
        priority="High",
        preconditions="Order qty >= 10; order pending approval",
        steps="1) Open Admin Order Details\n2) Try setting Shipping Fee to a non-zero value\n3) Save",
        expected="Server rejects or forces shipping fee to 0; no bypass.",
    ),
]


COLUMNS = [
    "TC ID",
    "Area",
    "Title",
    "Priority",
    "Preconditions",
    "Steps",
    "Expected Result",
    "Status (Pass/Fail)",
    "Actual Result / Notes",
    "Evidence (link/path)",
]


def autosize_columns(ws):
    # Cap width so long text doesn't explode the sheet.
    max_width = 60
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            text = str(val)
            # consider first line only for width; wrapping handles the rest
            first_line = text.splitlines()[0] if "\n" in text else text
            max_len = max(max_len, len(first_line))
        width = min(max_width, max(12, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_xlsx(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    header_fill = PatternFill("solid", fgColor="1F2937")  # slate-ish
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(COLUMNS)
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wrap = Alignment(vertical="top", wrap_text=True)

    for tc in TEST_CASES:
        ws.append(
            [
                tc.id,
                tc.area,
                tc.title,
                tc.priority,
                tc.preconditions,
                tc.steps,
                tc.expected,
                "Not Run",
                "",
                "",
            ]
        )

    for row in range(2, ws.max_row + 1):
        for col in range(1, len(COLUMNS) + 1):
            ws.cell(row=row, column=col).alignment = wrap

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    autosize_columns(ws)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_csv(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        for tc in TEST_CASES:
            w.writerow(
                [
                    tc.id,
                    tc.area,
                    tc.title,
                    tc.priority,
                    tc.preconditions,
                    tc.steps,
                    tc.expected,
                    "Not Run",
                    "",
                    "",
                ]
            )


def main() -> None:
    out_dir = Path(__file__).resolve().parent
    write_xlsx(out_dir / "test-cases.xlsx")
    write_csv(out_dir / "test-cases.csv")
    print("Wrote:")
    print(f"- {out_dir / 'test-cases.xlsx'}")
    print(f"- {out_dir / 'test-cases.csv'}")


if __name__ == "__main__":
    main()
